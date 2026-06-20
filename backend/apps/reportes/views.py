from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.utils import timezone
from apps.asistencia.models import RegistroAsistencia
import datetime
import io


def obtener_registros(inicio, fin):
    return RegistroAsistencia.objects.filter(
        timestamp__range=(inicio, fin)
    ).select_related('empleado', 'empleado__departamento').order_by('timestamp')


def construir_resumen(registros):
    tz = timezone.get_current_timezone()
    resumen = {}
    for reg in registros:
        ts_local = timezone.localtime(reg.timestamp, tz)
        clave    = (reg.empleado.id, ts_local.date())
        if clave not in resumen:
            departamento_nombre = reg.empleado.departamento.nombre if reg.empleado.departamento else None
            if isinstance(departamento_nombre, str) and departamento_nombre.strip().lower() == 'personal de limpieza':
                departamento_nombre = 'No aplica'

            resumen[clave] = {
                'nombre':       reg.empleado.nombre_completo,
                'cedula':       reg.empleado.cedula,
                'departamento': departamento_nombre,
                'tipo':         reg.empleado.tipo,
                'fecha':        ts_local.date().strftime('%d/%m/%Y'),
                'entrada':      '',
                'salida':       '',
                'retardo':      False,
                'minutos':      0,
            }
        if reg.tipo == 'entrada':
            resumen[clave]['entrada'] = ts_local.strftime('%I:%M %p')
            resumen[clave]['retardo'] = reg.es_retardo
            resumen[clave]['minutos'] = reg.minutos_retardo
        elif reg.tipo == 'salida':
            resumen[clave]['salida'] = ts_local.strftime('%I:%M %p')
    return list(resumen.values())


def rango_desde_params(request, modo):
    """
    Calcula inicio y fin según los parámetros de la petición.
    modo: 'dia' | 'semana' | 'mes'
    """
    tz   = timezone.get_current_timezone()
    hoy  = timezone.localtime(timezone.now()).date()

    if modo == 'dia':
        fecha_str = request.query_params.get('fecha', str(hoy))
        try:
            fecha = datetime.date.fromisoformat(fecha_str)
        except ValueError:
            fecha = hoy
        inicio = timezone.make_aware(datetime.datetime.combine(fecha, datetime.time.min), tz)
        fin    = timezone.make_aware(datetime.datetime.combine(fecha, datetime.time.max), tz)
        label  = fecha.strftime('%d/%m/%Y')

    elif modo == 'semana':
        fecha_str = request.query_params.get('fecha', str(hoy))
        try:
            fecha = datetime.date.fromisoformat(fecha_str)
        except ValueError:
            fecha = hoy
        # Lunes de esa semana
        lunes  = fecha - datetime.timedelta(days=fecha.weekday())
        domingo = lunes + datetime.timedelta(days=6)
        inicio = timezone.make_aware(datetime.datetime.combine(lunes,   datetime.time.min), tz)
        fin    = timezone.make_aware(datetime.datetime.combine(domingo, datetime.time.max), tz)
        label  = f'{lunes.strftime("%d/%m/%Y")} al {domingo.strftime("%d/%m/%Y")}'

    elif modo == 'mes':
        anio_str = request.query_params.get('anio', str(hoy.year))
        mes_str  = request.query_params.get('mes',  str(hoy.month))
        try:
            anio = int(anio_str)
            mes  = int(mes_str)
        except ValueError:
            anio, mes = hoy.year, hoy.month
        primer_dia = datetime.date(anio, mes, 1)
        if mes == 12:
            ultimo_dia = datetime.date(anio + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            ultimo_dia = datetime.date(anio, mes + 1, 1) - datetime.timedelta(days=1)
        inicio = timezone.make_aware(datetime.datetime.combine(primer_dia, datetime.time.min), tz)
        fin    = timezone.make_aware(datetime.datetime.combine(ultimo_dia, datetime.time.max), tz)
        label  = primer_dia.strftime('%B %Y')

    return inicio, fin, label


def generar_pdf(titulo, subtitulo, filas):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elems  = []

    elems.append(Paragraph('UPTAIET — Sede Rubio', styles['Title']))
    elems.append(Paragraph(titulo,    styles['Heading2']))
    elems.append(Paragraph(subtitulo, styles['Normal']))
    elems.append(Spacer(1, 16))

    datos = [['Empleado', 'Cedula', 'Departamento', 'Tipo', 'Fecha', 'Entrada', 'Salida', 'Estado']]
    for f in filas:
        estado = f'Retraso {f["minutos"]}min' if f['retardo'] else 'A tiempo'
        datos.append([
            f['nombre'], f['cedula'], f['departamento'], f['tipo'],
            f['fecha'],  f['entrada'], f['salida'], estado
        ])

    if len(datos) == 1:
        datos.append(['Sin registros en este periodo', '', '', '', '', '', '', ''])

    tabla = Table(datos, colWidths=[140, 75, 110, 80, 65, 65, 65, 90])
    tabla.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elems.append(tabla)
    elems.append(Spacer(1, 12))

    total    = len(filas)
    retrasos = sum(1 for f in filas if f['retardo'])
    tz_now   = timezone.localtime(timezone.now())
    elems.append(Paragraph(
        f'Total registros: {total} | Retrasos: {retrasos} | '
        f'Generado: {tz_now.strftime("%d/%m/%Y %I:%M %p")}',
        styles['Normal']
    ))

    doc.build(elems)
    buffer.seek(0)
    return buffer


def generar_excel(titulo, filas):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    ws.merge_cells('A1:H1')
    ws['A1']           = titulo
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', fgColor='1e3a5f')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Empleado', 'Cedula', 'Departamento', 'Tipo Empleado', 'Fecha', 'Entrada', 'Salida', 'Estado']
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='2d5a9e')
        cell.alignment = Alignment(horizontal='center')

    for row, f in enumerate(filas, 3):
        ws.cell(row=row, column=1, value=f['nombre'])
        ws.cell(row=row, column=2, value=f['cedula'])
        ws.cell(row=row, column=3, value=f['departamento'])
        ws.cell(row=row, column=4, value=f['tipo'])
        ws.cell(row=row, column=5, value=f['fecha'])
        ws.cell(row=row, column=6, value=f['entrada'])
        ws.cell(row=row, column=7, value=f['salida'])
        ws.cell(row=row, column=8, value=f'Retraso {f["minutos"]}min' if f['retardo'] else 'A tiempo')

    for col in range(1, 9):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Vistas ────────────────────────────────────────────────

class ReporteDiarioPDFView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inicio, fin, label = rango_desde_params(request, 'dia')
        filas  = construir_resumen(obtener_registros(inicio, fin))
        buffer = generar_pdf(f'Reporte Diario — {label}', f'Asistencia del {label}', filas)
        res = HttpResponse(buffer, content_type='application/pdf')
        res['Content-Disposition'] = f'attachment; filename="diario_{label.replace("/","-")}.pdf"'
        return res

class ReporteSemanalPDFView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inicio, fin, label = rango_desde_params(request, 'semana')
        filas  = construir_resumen(obtener_registros(inicio, fin))
        buffer = generar_pdf(f'Reporte Semanal — {label}', 'Asistencia de la semana', filas)
        res = HttpResponse(buffer, content_type='application/pdf')
        res['Content-Disposition'] = f'attachment; filename="semanal_{label.replace("/","-")}.pdf"'
        return res

class ReporteMensualPDFView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inicio, fin, label = rango_desde_params(request, 'mes')
        filas  = construir_resumen(obtener_registros(inicio, fin))
        buffer = generar_pdf(f'Reporte Mensual — {label}', f'Asistencia de {label}', filas)
        res = HttpResponse(buffer, content_type='application/pdf')
        res['Content-Disposition'] = f'attachment; filename="mensual_{label.replace(" ","_")}.pdf"'
        return res

class ReporteDiarioExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inicio, fin, label = rango_desde_params(request, 'dia')
        filas  = construir_resumen(obtener_registros(inicio, fin))
        buffer = generar_excel(f'UPTAIET — Asistencia Diaria {label}', filas)
        res = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename="diario_{label.replace("/","-")}.xlsx"'
        return res

class ReporteSemanalExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inicio, fin, label = rango_desde_params(request, 'semana')
        filas  = construir_resumen(obtener_registros(inicio, fin))
        buffer = generar_excel(f'UPTAIET — Asistencia Semanal {label}', filas)
        res = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename="semanal_{label.replace("/","-")}.xlsx"'
        return res

class ReporteMensualExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        inicio, fin, label = rango_desde_params(request, 'mes')
        filas  = construir_resumen(obtener_registros(inicio, fin))
        buffer = generar_excel(f'UPTAIET — Asistencia Mensual {label}', filas)
        res = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename="mensual_{label.replace(" ","_")}.xlsx"'
        return res